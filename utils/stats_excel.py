import os
from decimal import Decimal
from openpyxl import Workbook, load_workbook

STATS_FILE = "stats.xlsx"

def load_or_create_workbook():
    if not os.path.exists(STATS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Stats"
        ws.append(["NAME", "SYMBOL", "TOTAL_VOLUME_USD", "LONG_COUNT", "SHORT_COUNT"])
        wb.save(STATS_FILE)
    return load_workbook(STATS_FILE)

def update_stats_excel(name: str, symbol: str, qty: Decimal, side: str, mark_price: Decimal):
    wb = load_or_create_workbook()
    ws = wb["Stats"]

    usd_volume = float(qty * mark_price)

    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == name and row[1].value == symbol:
            row[2].value = float(row[2].value or 0) + usd_volume
            if side == "BUY":
                row[3].value = int(row[3].value or 0) + 1
            elif side == "SELL":
                row[4].value = int(row[4].value or 0) + 1
            wb.save(STATS_FILE)
            return

    long_count = 1 if side == "BUY" else 0
    short_count = 1 if side == "SELL" else 0
    ws.append([name, symbol, usd_volume, long_count, short_count])
    wb.save(STATS_FILE)